
"""
Data Quality Check Script
Version: 1.0
Date: 2025-10-31

This script validates merged clean data against QC criteria.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import glob
import os
import re

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

US_STATES = [
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
    "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
    "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"
]


def find_latest_clean_file():
    """Find the most recent MERGE_CLEAN_EXCEL_FILES_AUTO_*.xlsx file."""
    pattern = "MERGE_CLEAN_EXCEL_FILES_AUTO_*.xlsx"
    files = glob.glob(pattern)

    if not files:
        raise FileNotFoundError(f"No files matching pattern '{pattern}' found in current directory")

    latest_file = max(files, key=os.path.getmtime)
    print(f"Found latest clean file: {latest_file}")
    return latest_file


def load_qc_criteria():
    """Load QC criteria from CELL QC CRITERIA.xlsx."""
    criteria_file = "CELL QC CRITERIA.xlsx"

    if not os.path.exists(criteria_file):
        raise FileNotFoundError(f"QC criteria file not found: {criteria_file}")

    # Load both sheets
    full_list = pd.read_excel(criteria_file, sheet_name="FULL LIST")
    phase_equivalent = pd.read_excel(criteria_file, sheet_name="phase equivalent")

    print(f"Loaded {len(full_list)} column criteria from FULL LIST sheet")
    print(f"Loaded {len(phase_equivalent)} phase mappings from phase equivalent sheet")

    # Create criteria dictionary: column_name -> validation_rule
    criteria_dict = {}
    for _, row in full_list.iterrows():
        col_name = row["COLUMN NAME"]
        valid_rule = row["VALID"]
        if pd.notna(valid_rule) and str(valid_rule).strip():
            criteria_dict[col_name] = str(valid_rule).strip()

    # Create phase mapping dictionary: PHASES -> Phase_CALC
    phase_map = {}
    for _, row in phase_equivalent.iterrows():
        phases = row["PHASES"]
        phase_calc = row["Phase_CALC"]
        if pd.notna(phases) and pd.notna(phase_calc):
            phase_map[str(phases).strip()] = str(phase_calc).strip()

    return criteria_dict, phase_map


def check_cell(value, rule, col_name, row_data):
    """
    Validate a single cell value against its QC rule.

    Returns: (is_valid, error_message)
    """
    # Handle empty/null values first
    is_empty = pd.isna(value) or str(value).strip() == ""

    # Parse the rule
    rule_upper = rule.upper()

    # Check for source-specific "only empty in" requirements
    source = row_data.get("SOURCE", "")

    # Handle "only empty in CAM_Run_tracker" validation
    if "only empty in CAM" in rule.lower() or "only empty in CAM_Run_tracker" in rule:
        if source != "CAM_Run_Tracker" and is_empty:
            return False, f"Required field (empty only allowed in CAM_Run_Tracker)"
        return True, None

    # Handle "only empty in POG" validation
    if "only empty in POG" in rule:
        if source not in ["POG_MM_Usage", "POG_CAM_Usage"] and is_empty:
            return False, f"Required field (empty only allowed in POG files)"
        return True, None

    # Handle "only empty in CAM_run_tracker, POG_MM_Usage, POG_CAM_Usage"
    if "only empty in CAM_run_tracker, POG_MM_Usage, POG_CAM_Usage" in rule:
        if source not in ["CAM_Run_Tracker", "POG_MM_Usage", "POG_CAM_Usage"] and is_empty:
            return False, f"Required field (empty only allowed in CAM/POG files)"
        return True, None

    # Handle "NON-BLANK in Motor KPI and CAM Run tracker, empty in POG files"
    if "Motor KPI and CAM Run tracker, empty in POG" in rule:
        if source in ["Motor_KPI", "CAM_Run_Tracker"] and is_empty:
            return False, f"Required for Motor_KPI and CAM_Run_Tracker"
        return True, None

    # Check for NON-BLANK / NON EMPTY requirements
    if "NON-BLANK" in rule_upper or "NON EMPTY" in rule_upper:
        # Check for conditional requirements
        if "IF" in rule_upper:
            # Handle CUR-related fields: required if CUR in Phase_CALC (only for Motor_KPI)
            if "CUR" in rule_upper and "PHASE_CALC" in rule_upper:
                # Check if this is a Motor_KPI-only rule
                if "only for Motor_KPI" in rule or "ONLY FOR MOTOR_KPI" in rule_upper:
                    # Only validate for Motor_KPI rows
                    if source != "Motor_KPI":
                        return True, None

                phase_calc = row_data.get("Phase_CALC", "")
                if pd.notna(phase_calc) and "CUR" in str(phase_calc).upper():
                    if is_empty:
                        return False, "Required when CUR in Phase_CALC"
                return True, None

            # Handle REPORTED_AS rule: required if INCIDENT_NUM is non-blank
            elif col_name == "REPORTED_AS":
                incident_num = row_data.get("INCIDENT_NUM", "")
                if pd.notna(incident_num) and str(incident_num).strip() != "":
                    if is_empty:
                        return False, "Required when INCIDENT_NUM is filled"
                return True, None
        else:
            # Unconditional non-blank requirement
            if is_empty:
                return False, "Required field is empty"

    # If value is empty and not caught by above rules, it's OK
    if is_empty:
        return True, None

    # Check numeric range validation (e.g., "<30000", "<600")
    if rule.startswith("<"):
        try:
            threshold = float(rule[1:].strip())
            cell_value = float(value)
            if cell_value >= threshold:
                return False, f"Value {cell_value} exceeds limit {threshold}"
        except (ValueError, TypeError):
            return False, f"Expected numeric value for range check"

    # Check for list validation (comma-separated allowed values)
    if "," in rule and not rule_upper.startswith("NON"):
        allowed_values = [v.strip() for v in rule.split(",")]
        cell_value_str = str(value).strip()
        if cell_value_str not in allowed_values:
            return False, f"Value '{cell_value_str}' not in allowed list"

    # Check for "Number" type validation
    if rule.upper() == "NUMBER":
        try:
            float(value)
        except (ValueError, TypeError):
            return False, f"Expected numeric value"

    # Check for state validation
    if col_name == "STATE":
        state_str = str(value).strip().upper()
        if state_str not in US_STATES:
            return False, f"Invalid state code: {state_str}"

    # Special handling for SOURCE column validation
    if col_name == "SOURCE":
        valid_sources = ["Motor_KPI", "CAM_Run_Tracker", "POG_MM_Usage", "POG_CAM_Usage"]
        cell_value_str = str(value).strip()
        if cell_value_str not in valid_sources:
            return False, f"Value '{cell_value_str}' not in allowed source list"

    # Check Phase_CALC validation
    if col_name == "Phase_CALC":
        # This would need the phase_map, but checking is done separately
        # in validate_data function
        pass

    return True, None


def validate_data(df, criteria_dict, phase_map):
    """
    Validate entire dataframe against QC criteria.

    Returns: dict of issues {(row_idx, col_name): error_message}
    """
    issues_by_cell = {}

    print(f"\nValidating {len(df)} rows against {len(criteria_dict)} criteria...")

    for row_idx in range(len(df)):
        row_data = df.iloc[row_idx]

        # Check each column that has criteria
        for col_name, rule in criteria_dict.items():
            if col_name not in df.columns:
                continue

            value = row_data[col_name]

            # Special handling for Phase_CALC validation
            if col_name == "Phase_CALC":
                phases_value = row_data.get("PHASES", "")
                if pd.notna(phases_value):
                    phases_str = str(phases_value).strip()
                    expected_phase_calc = phase_map.get(phases_str)
                    if expected_phase_calc:
                        actual_phase_calc = str(value).strip() if pd.notna(value) else ""
                        if actual_phase_calc != expected_phase_calc:
                            issues_by_cell[(row_idx, col_name)] = f"Expected '{expected_phase_calc}' for PHASES='{phases_str}'"
                continue

            # Standard validation
            is_valid, error_msg = check_cell(value, rule, col_name, row_data)
            if not is_valid:
                issues_by_cell[(row_idx, col_name)] = error_msg

    print(f"Found {len(issues_by_cell)} cell issues")
    return issues_by_cell


def apply_qc_flag(df, issues_by_cell):
    """Add QC_FLAG column: 1 if row has issues, 0 if clean."""
    qc_flags = []

    for row_idx in range(len(df)):
        # Check if this row has any issues
        has_issue = any(cell_key[0] == row_idx for cell_key in issues_by_cell.keys())
        qc_flags.append(1 if has_issue else 0)

    df["QC_FLAG"] = qc_flags

    rows_with_issues = sum(qc_flags)
    print(f"QC_FLAG column added: {rows_with_issues} rows with issues, {len(df) - rows_with_issues} clean rows")


def highlight_issues_in_excel(output_file, df, issues_by_cell):
    """Apply yellow highlighting to cells with issues."""
    print(f"\nApplying yellow highlighting to {len(issues_by_cell)} cells...")

    # Convert DATE_IN and DATE_OUT to date-only format (remove time)
    date_columns = ['DATE_IN', 'DATE_OUT']
    for col in date_columns:
        if col in df.columns:
            # Convert to datetime if not already, then format as date only
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

    # First, save the dataframe to Excel
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Now open with openpyxl to apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Apply yellow fill to each issue cell
    for (row_idx, col_name), error_msg in issues_by_cell.items():
        # Excel rows are 1-indexed, and we have a header row
        excel_row = row_idx + 2  # +1 for 0-index, +1 for header

        # Find column index
        col_idx = df.columns.get_loc(col_name) + 1  # Excel is 1-indexed

        # Apply yellow fill
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.fill = YELLOW_FILL

    wb.save(output_file)
    print(f"Highlighting applied and saved to: {output_file}")


def main():
    """Main execution function."""
    print("=" * 60)
    print("QC Data Quality Check Script")
    print("=" * 60)

    try:
        # Find latest clean file
        input_file = find_latest_clean_file()

        # Load QC criteria
        criteria_dict, phase_map = load_qc_criteria()

        # Load data
        print(f"\nLoading data from: {input_file}")
        df = pd.read_excel(input_file)
        print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

        # Validate data
        issues_by_cell = validate_data(df, criteria_dict, phase_map)

        # Add QC_FLAG column
        apply_qc_flag(df, issues_by_cell)

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"MERGE_CLEAN_QC_{timestamp}.xlsx"

        # Highlight issues and save
        highlight_issues_in_excel(output_file, df, issues_by_cell)

        # Summary
        print("\n" + "=" * 60)
        print("QC SUMMARY")
        print("=" * 60)
        print(f"Input file: {input_file}")
        print(f"Output file: {output_file}")
        print(f"Total rows: {len(df)}")
        print(f"Total cells checked: {len(df) * len(criteria_dict)}")
        print(f"Issues found: {len(issues_by_cell)} cells")
        print(f"Rows with issues: {df['QC_FLAG'].sum()}")
        print(f"Clean rows: {len(df) - df['QC_FLAG'].sum()}")
        print("=" * 60)

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
